import os
import discord
import openpyxl

from discord.ext import commands
from datetime import datetime
from dotenv import load_dotenv

from dropdown import ReportView
from game import Player, Game, Team, TeamGameStatus
from game_api_client import GameAPI

load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
API_KEY = os.getenv('API_KEY')
DEV_MOD = os.getenv('DEV_MODE')
BOT_ID = int(os.getenv('BOT_ID'))
CHANNEL_ID_2VI2 = int(os.getenv('CHANNEL_ID_2VI2'))
CHANNEL_ID_RANKED = int(os.getenv('CHANNEL_ID_RANKED'))
REPORTS_IDS = [int(x) for x in os.getenv('REPORTS_IDS').split(",")]

intents = discord.Intents.all()
bot = commands.Bot(command_prefix='!', intents=intents)


def parse_message_to_game(date, report, players):
    tournament_name = 'Season 2 - Tournament 2x2'
    players = [
        Player(str(players[0].id), players[0].name),
        Player(str(players[1].id), players[1].name),
        Player(str(players[2].id), players[2].name),
        Player(str(players[3].id), players[3].name)
    ]

    team1 = get_team("1", report, players)
    team2 = get_team("2", report, players)

    return Game(date, tournament_name, [team1, team2])


def get_team(team_name, report, players):
    team_civilizations = {}
    for line in report.split("\n"):
        if "T" + team_name in line:
            tokens = line.split()
            discord_id = tokens[3][2:-1]
            civilization = tokens[4]
            team_civilizations[discord_id] = civilization

    team_players = []
    for player in players:
        if str(player.discordId) in team_civilizations:
            player.set_civilization(team_civilizations[str(player.discordId)])
            team_players.append(player)

    status = TeamGameStatus
    for line in report.split("\n"):
        if "Winning team:" in line:
            winning_team_str = line.split(": ")[1]
            if winning_team_str == "Team " + team_name:
                status = TeamGameStatus.WON
            else:
                status = TeamGameStatus.LOST

    banned_civs = report.split("\n")[3].split(": ")[1].split(",")

    team = Team("Team " + team_name, team_players, [banned_civs[int(team_name) - 1].replace(' ', '')], status)

    return team


def excel_writer(reports):
    xname = "reports.xlsx"
    xwb = openpyxl.load_workbook(xname)
    xws = xwb.active
    last_row = xws.max_row
    xws.cell(row=last_row + 1, column=1, value=last_row)
    xws.cell(row=last_row + 1, column=2, value=reports[0])
    xws.cell(row=last_row + 1, column=3, value=reports[1])
    xws.cell(row=last_row + 1, column=4, value=reports[2])
    xwb.save(xname)


@bot.command()
async def report(ctx, member1: discord.Member, member2: discord.Member,
                 member3: discord.Member, member4: discord.Member):
    report_info = [[ctx.author], [member1, member2, member3, member4]]
    view = ReportView(ctx, report_info)

    await ctx.send('Please select an option:', view=view)


@bot.event
async def on_reaction_add(reaction, user):
    message = reaction.message
    if message.channel.id == CHANNEL_ID_2VI2:
        if message.author.id == BOT_ID:

            if reaction.emoji == '✅' and user.id in REPORTS_IDS:
                channel = message.channel
                async for mes in channel.history(limit=1, before=message):
                    await mes.add_reaction('✅')
                await message.delete()

            players = message.mentions
            if reaction.emoji == '👍':
                if user.id in [player.id for player in players]:
                    players = [player for player in players if player.id != user.id]
                    if any(players):
                        await message.edit(
                            content=f"Players pending confirmation: "
                                    f"{','.join([player.mention for player in players])}"
                        )
                    else:
                        channel = message.channel
                        async for mes in channel.history(limit=1, before=message):
                            await mes.add_reaction('✅')
                        await message.delete()

            if reaction.emoji == '✅':
                if user.id == BOT_ID:

                    now = datetime.now()
                    now = now.strftime("%Y-%m-%d %H:%M:%S")

                    players_names = message.mentions
                    player_names = [player.name for player in players_names]
                    players_names = ', '.join(player_names)
                    games_reports = [now, message.content, players_names]
                    excel_writer(games_reports)

                    lines = message.content.splitlines()
                    lines[0] = lines[0].replace("GameType 2vi2", "GameType: Teamers 2vi2")
                    lines[4] = lines[4].replace("T1 slot 1: ", "").replace("_", "").replace("-", "")
                    lines[5] = lines[5].replace("T2 slot 2: ", "").replace("_", "").replace("-", "")
                    lines[6] = lines[6].replace("T2 slot 3: ", "").replace("_", "").replace("-", "")
                    lines[7] = lines[7].replace("T1 slot 4: ", "").replace("_", "").replace("-", "")

                    if "Team 1" in message.content:
                        selected_lines = [lines[0], lines[2], lines[4], lines[7], lines[5], lines[6]]
                    else:
                        selected_lines = [lines[0], lines[2], lines[5], lines[6], lines[4], lines[7]]

                    selected_lines[2] = f'1. {selected_lines[2]}'
                    selected_lines[3] = f'1. {selected_lines[3]}'
                    selected_lines[4] = f'\n2. {selected_lines[4]}'
                    selected_lines[5] = f'2. {selected_lines[5]}'

                    mod_channel = bot.get_channel(CHANNEL_ID_RANKED)
                    mod_report = '\n'.join(selected_lines)
                    await mod_channel.send(mod_report)

                    game = parse_message_to_game(now, message.content, players)
                    await post_game_to_civplays(game)


async def post_game_to_civplays(game):
    gameApi = GameAPI(API_KEY, dev_mode=DEV_MOD == 'True')
    gameApi.create_game(game)


@bot.command()
async def game_reports(ctx):
    if ctx.author.id in REPORTS_IDS:
        user = ctx.author
        file = discord.File("reports.xlsx")
        await user.send(file=file)

bot.run(TOKEN)
